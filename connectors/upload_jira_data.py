#!/usr/bin/env python3
"""
connectors/upload_jira_data.py
==============================
Reads the "Jira Tickets" sheet from the Release Management Platform Excel
file and bulk-uploads all data into Jira Cloud.

Steps performed:
  1. Ensure Fix Versions exist in each project.
  2. Create all issues (Stories, Bugs, Tasks, Test Executions).
  3. Transition issues to their target status (Done, In Progress, etc.).
  4. Create issue links (blocks, depends-on, relates-to, test-coverage).

Credentials are read from the project .env file automatically —
the same credentials used by jira_connector.py and sync.py.

Prerequisites:
  pip install openpyxl requests python-dotenv

Usage:
  python connectors/upload_jira_data.py --file "Release_Management_Platform_Sample_Data (1).xlsx"
  python connectors/upload_jira_data.py --file data.xlsx --dry-run
  python connectors/upload_jira_data.py --file data.xlsx --skip-transitions --skip-links
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from pathlib import Path

# ── Load .env from project root ────────────────────────────────────────────────
_ROOT = Path(__file__).parent.parent
try:
    from dotenv import load_dotenv
    load_dotenv(_ROOT / ".env")
except ImportError:
    pass  # dotenv optional — env vars can be set directly

import requests
from requests.auth import HTTPBasicAuth

# ── Credentials — reads from .env (same keys as jira_connector.py) ─────────────
JIRA_BASE_URL   = os.getenv("ATLASSIAN_DOMAIN", "").rstrip("/")
JIRA_USER_EMAIL = os.getenv("ATLASSIAN_EMAIL", "")
JIRA_API_TOKEN  = os.getenv("ATLASSIAN_API_TOKEN", "")
DRY_RUN         = os.getenv("DRY_RUN", "false").lower() == "true"

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("jira-uploader")

# ── Fix-version metadata per project ──────────────────────────────────────────
# Maps each Jira project key to its release version details.
# These match the Release Management Platform sample data guide.
FIX_VERSIONS = {
    "ATLAS":  {
        "name":        "v3.2 — May 2026 Release",
        "releaseDate": "2026-05-16",
        "description": "Payment service upgrade with PCI compliance fixes",
    },
    "BEACON": {
        "name":        "v2.0 — May 2026 Release",
        "releaseDate": "2026-05-16",
        "description": "Customer portal complete redesign",
    },
    "CIPHER": {
        "name":        "v1.5 — May 2026 Release",
        "releaseDate": "2026-05-16",
        "description": "API gateway security hardening",
    },
    "DELTA":  {
        "name":        "v4.1 — May 2026 Release",
        "releaseDate": "2026-05-16",
        "description": "Reporting dashboard enhancements",
    },
}

# ── Issue type mapping ─────────────────────────────────────────────────────────
# Spreadsheet type name → Jira issue type name.
# Adjust right-hand values to match your Jira project configuration.
ISSUE_TYPE_MAP = {
    "Story":          "Story",
    "Bug":            "Bug",
    "Task":           "Task",
    "Test Execution": "Test Execution",   # Requires Zephyr/Xray or custom type
    "Epic":           "Epic",
}
TEST_EXECUTION_FALLBACK = "Task"   # used if "Test Execution" type doesn't exist

# ── Priority mapping ───────────────────────────────────────────────────────────
PRIORITY_MAP = {
    "Blocker":  "Highest",
    "Critical": "High",
    "High":     "High",
    "Medium":   "Medium",
    "Low":      "Low",
}

# ── Status → workflow transition names ────────────────────────────────────────
# Jira requires *transitions* rather than direct status writes.
# Each list is the sequence of transition names to apply in order.
# IMPORTANT: these names must match your Jira workflow configuration.
STATUS_TRANSITIONS = {
    "Open":        [],
    "To Do":       [],
    "In Progress": ["In Progress"],   # Jira Cloud transition name
    "Done":        ["Done"],          # Jira Cloud transition name
    "Passed":      ["Done"],
    "Not Started": [],
}

# ── Link type mapping ──────────────────────────────────────────────────────────
# Keyword found in spreadsheet "Links" column → Jira link type + direction.
LINK_TYPE_MAP = {
    "blocks":        {"type": "Blocks",   "direction": "outward"},
    "is blocked by": {"type": "Blocks",   "direction": "inward"},
    "depends on":    {"type": "Blocks",   "direction": "inward"},
    "relates to":    {"type": "Relates",  "direction": "outward"},
    "test":          {"type": "Relates",  "direction": "outward"},  # "Test" not in standard Jira
    "story":         {"type": "Relates",  "direction": "inward"},
}


# ── Jira REST API client ───────────────────────────────────────────────────────

class JiraClient:
    """
    Thin wrapper around the Jira Cloud REST API v3.

    Uses the same credential format as jira_connector.py:
      ATLASSIAN_DOMAIN  / ATLASSIAN_EMAIL / ATLASSIAN_API_TOKEN
    """

    def __init__(self, base_url: str, email: str, token: str, dry_run: bool = False):
        self.base_url = base_url.rstrip("/")
        self.auth     = HTTPBasicAuth(email, token)
        self.headers  = {
            "Content-Type": "application/json",
            "Accept":       "application/json",
        }
        self.dry_run  = dry_run

        # Internal caches to avoid redundant API calls
        self._version_cache:    dict[str, dict[str, str]] = {}
        self._issue_type_cache: dict[str, list[str]]      = {}
        self.key_to_id:         dict[str, str]            = {}
        # Maps Excel key (e.g. ATLAS-101) -> actual Jira key (e.g. ATLAS-1)
        self.excel_to_actual:   dict[str, str]            = {}

    def _request(self, method: str, path: str, **kwargs):
        """
        Make one authenticated HTTP request to Jira REST API v3.
        Handles rate limiting (429) with automatic retry.
        Logs and raises on 4xx/5xx errors.
        """
        url = f"{self.base_url}/rest/api/3{path}"
        if self.dry_run and method.upper() != "GET":
            log.info("[DRY-RUN] %s %s", method, url)
            return None
        resp = requests.request(method, url, auth=self.auth,
                                headers=self.headers, timeout=20, **kwargs)
        if resp.status_code == 429:
            # Jira rate limit — wait and retry
            wait = int(resp.headers.get("Retry-After", "5"))
            log.warning("Rate-limited. Waiting %ds ...", wait)
            time.sleep(wait)
            return self._request(method, path, **kwargs)
        if resp.status_code >= 400:
            log.error("HTTP %d  %s %s\n  %s",
                      resp.status_code, method, url, resp.text[:400])
        resp.raise_for_status()
        return resp.json() if resp.text else None

    def get(self, path, **kw):  return self._request("GET",  path, **kw)
    def post(self, path, data): return self._request("POST", path, json=data)
    def put(self,  path, data): return self._request("PUT",  path, json=data)

    # ── Project metadata ───────────────────────────────────────────────────────

    def get_project(self, key: str) -> dict | None:
        """Return project metadata dict, or None if project not found."""
        try:
            return self.get(f"/project/{key}")
        except requests.HTTPError:
            return None

    def get_issue_types(self, project_key: str) -> list[str]:
        """
        Return available issue-type names for a project.
        Uses /issue/createmeta (v3 format).
        """
        if self.dry_run:
            return ["Story", "Bug", "Task", "Epic"]   # safe defaults for dry-run

        if project_key not in self._issue_type_cache:
            meta  = self.get(f"/issue/createmeta?projectKeys={project_key}"
                             "&expand=projects.issuetypes")
            names = []
            for proj in (meta or {}).get("projects", []):
                for it in proj.get("issuetypes", []):
                    names.append(it["name"])
            self._issue_type_cache[project_key] = names
        return self._issue_type_cache[project_key]

    # ── Fix versions ───────────────────────────────────────────────────────────

    def ensure_fix_version(self, project_key: str, version_info: dict) -> str:
        """
        Create a fix version in the project if it does not already exist.
        Returns the version ID.
        Uses a local cache to avoid duplicate GET calls.
        """
        if self.dry_run:
            log.info("[DRY-RUN] Would ensure fix version '%s' in %s",
                     version_info["name"], project_key)
            return "dry-run-id"

        if project_key not in self._version_cache:
            existing = self.get(f"/project/{project_key}/versions") or []
            self._version_cache[project_key] = {v["name"]: v["id"] for v in existing}

        cache = self._version_cache[project_key]
        if version_info["name"] in cache:
            log.info("Fix version '%s' already exists in %s",
                     version_info["name"], project_key)
            return cache[version_info["name"]]

        proj    = self.get_project(project_key)
        payload = {
            "name":        version_info["name"],
            "description": version_info.get("description", ""),
            "releaseDate": version_info.get("releaseDate"),
            "released":    False,
            "projectId":   int(proj["id"]),
        }
        result = self.post("/version", payload)
        vid    = result["id"] if result else "dry-run-id"
        cache[version_info["name"]] = vid
        log.info("Created fix version '%s' in %s (id=%s)",
                 version_info["name"], project_key, vid)
        return vid

    # ── Issue creation ─────────────────────────────────────────────────────────

    def resolve_issue_type(self, project_key: str, desired: str) -> str:
        """
        Map a spreadsheet type name to a valid Jira issue type for the project.
        Falls back to Task if Test Execution is not available.
        """
        available = self.get_issue_types(project_key)
        mapped    = ISSUE_TYPE_MAP.get(desired, desired)
        if mapped in available:
            return mapped
        if desired == "Test Execution" and TEST_EXECUTION_FALLBACK in available:
            log.warning("'Test Execution' not in %s — using '%s'",
                        project_key, TEST_EXECUTION_FALLBACK)
            return TEST_EXECUTION_FALLBACK
        return mapped

    def create_issue(self, project_key: str, row: dict) -> str | None:
        """
        Create one Jira issue from a spreadsheet row dict.

        Parameters
        ----------
        project_key : str   The Jira project key (e.g. "ATLAS").
        row         : dict  Keys: Ticket Key, Type, Summary, Priority,
                            Status, Assignee, Links, Milestone.

        Returns
        -------
        str   The created issue key (e.g. "ATLAS-101"), or None on failure.
        """
        issue_type       = self.resolve_issue_type(project_key, row["Type"])
        fix_version_name = FIX_VERSIONS.get(project_key, {}).get("name", "")

        fields: dict = {
            "project":   {"key": project_key},
            "issuetype": {"name": issue_type},
            "summary":   row["Summary"],
            "priority":  {"name": PRIORITY_MAP.get(row.get("Priority", "Medium"),
                                                    row.get("Priority", "Medium"))},
        }
        if fix_version_name:
            fields["fixVersions"] = [{"name": fix_version_name}]

        # Assignee — skip placeholder names like "Dev1", "Dev2"
        assignee = row.get("Assignee", "")
        if assignee and not re.match(r"^(Dev|QA)\d+$", assignee, re.I):
            # Jira Cloud uses accountId; Server/DC uses name
            fields["assignee"] = {"name": assignee}

        # Label from milestone column
        milestone = row.get("Milestone", "")
        if milestone:
            label = re.sub(r"[^A-Za-z0-9_-]", "_", milestone)
            fields["labels"] = [label]

        log.info("Creating %s %s: %s", issue_type, row["Ticket Key"], row["Summary"])
        excel_key = row["Ticket Key"]
        try:
            result = self.post("/issue", {"fields": fields})
            if result:
                actual_key = result["key"]
                self.key_to_id[actual_key] = result["id"]
                # Store Excel key → actual Jira key so transitions and links work
                self.excel_to_actual[excel_key] = actual_key
                log.info("  → Created %s (Excel: %s)", actual_key, excel_key)
                return actual_key
            else:
                # Dry-run: map Excel key to itself
                self.excel_to_actual[excel_key] = excel_key
                self.key_to_id[excel_key] = f"dry-{excel_key}"
                return excel_key
        except requests.HTTPError as exc:
            log.error("  ✗ Failed %s: %s", row["Ticket Key"], exc)
            return None

    # ── Status transitions ─────────────────────────────────────────────────────

    def get_transitions(self, issue_key: str) -> dict[str, str]:
        """Return {transition_name_lower: transition_id} for an issue."""
        if self.dry_run:
            return {}
        data = self.get(f"/issue/{issue_key}/transitions")
        return {t["name"].lower(): t["id"]
                for t in (data or {}).get("transitions", [])}

    def transition_issue(self, issue_key: str, target_status: str):
        """
        Move an issue through workflow steps to reach target_status.
        Applies transitions in sequence from STATUS_TRANSITIONS.
        Uses fuzzy matching if exact transition name not found.
        """
        steps = STATUS_TRANSITIONS.get(target_status, [])
        if not steps:
            return
        if self.dry_run:
            log.info("[DRY-RUN] Would transition %s → %s (steps: %s)",
                     issue_key, target_status, steps)
            return

        for step in steps:
            available = self.get_transitions(issue_key)
            tid = available.get(step.lower())
            if tid:
                self.post(f"/issue/{issue_key}/transitions",
                          {"transition": {"id": tid}})
                log.info("  → %s transitioned via '%s'", issue_key, step)
            else:
                # Fuzzy match — Jira workflow names vary by instance
                matched = False
                for name, tid2 in available.items():
                    if step.lower() in name or name in step.lower():
                        self.post(f"/issue/{issue_key}/transitions",
                                  {"transition": {"id": tid2}})
                        log.info("  → %s transitioned via '%s' (fuzzy: '%s')",
                                 issue_key, name, step)
                        matched = True
                        break
                if not matched:
                    log.warning("  ⚠ '%s' not found for %s. Available: %s",
                                step, issue_key, list(available.keys()))

    # ── Issue links ────────────────────────────────────────────────────────────

    def create_link(self, link_type: str, inward_key: str, outward_key: str):
        """Create a directional issue link between two tickets."""
        payload = {
            "type":         {"name": link_type},
            "inwardIssue":  {"key": inward_key},
            "outwardIssue": {"key": outward_key},
        }
        log.info("  → Link: %s —[%s]→ %s", inward_key, link_type, outward_key)
        try:
            self.post("/issueLink", payload)
        except requests.HTTPError as exc:
            log.error("  ✗ Link failed %s → %s: %s",
                      inward_key, outward_key, exc)

    def load_existing_keys(self, project_keys: list, excel_tickets: list):
        """
        Fetch all existing issues from Jira for the given projects and
        build the excel_to_actual mapping by matching summaries.

        Used when --skip-issues is passed to avoid re-creating tickets
        that were already uploaded. Matches each Excel ticket summary
        to the actual Jira key (e.g. ATLAS-101 summary → ATLAS-1).
        """
        for pkey in project_keys:
            log.info("  Loading existing issues from %s ...", pkey)
            start, page_size = 0, 100
            summary_to_key: dict[str, str] = {}
            while True:
                data = self.get(
                    f"/search/jql?jql=project={pkey}&fields=summary,key"
                    f"&maxResults={page_size}&startAt={start}"
                )
                if not data:
                    break
                issues = data.get("issues", [])
                for issue in issues:
                    summary = issue.get("fields", {}).get("summary", "").strip()
                    summary_to_key[summary] = issue["key"]
                if len(issues) < page_size:
                    break
                start += page_size
            log.info("    Found %d existing issues in %s", len(summary_to_key), pkey)

            # Match each Excel ticket to an existing Jira issue by summary
            for ticket in excel_tickets:
                if ticket.get("Project") != pkey:
                    continue
                summary    = ticket.get("Summary", "").strip()
                actual_key = summary_to_key.get(summary)
                if actual_key:
                    excel_key = ticket["Ticket Key"]
                    self.excel_to_actual[excel_key] = actual_key
                    self.key_to_id[actual_key]      = actual_key  # placeholder
                else:
                    log.warning("  No match found for: %s — %s",
                                ticket["Ticket Key"], summary[:60])


# ── Excel parser ───────────────────────────────────────────────────────────────

def read_jira_sheet(file_path: str) -> list[dict]:
    """
    Read the 'Jira Tickets' sheet from the Excel file.
    Returns a list of row dicts keyed by the header row.
    Skips rows where Ticket Key or Summary is empty.
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path, read_only=True)

    if "Jira Tickets" not in wb.sheetnames:
        log.error("Sheet 'Jira Tickets' not found. Available: %s", wb.sheetnames)
        sys.exit(1)

    ws      = wb["Jira Tickets"]
    rows    = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data    = []

    for row in rows[1:]:
        record = {headers[i]: (str(v).strip() if v is not None else "")
                  for i, v in enumerate(row)}
        # Skip empty rows
        if not record.get("Ticket Key") or not record.get("Summary"):
            continue
        data.append(record)

    wb.close()
    log.info("Read %d tickets from 'Jira Tickets' sheet.", len(data))
    return data


def parse_links(links_cell: str) -> list[dict]:
    """
    Parse the 'Links' column value into a list of structured link dicts.

    Supported formats:
      "Test: ATLAS-201"
      "Blocks: BEACON-105, Depends on: CIPHER-104"
      "Is blocked by: ATLAS-102"
      "—" or "" → empty list

    Returns:
      [{"keyword": "blocks", "target": "BEACON-105"}, ...]
    """
    if not links_cell or links_cell in ("—", "None", "N/A", ""):
        return []
    results  = []
    segments = [s.strip() for s in links_cell.split(",")]
    for seg in segments:
        m = re.match(r"^(.+?):\s*([A-Z]+-\d+)$", seg.strip(), re.IGNORECASE)
        if m:
            results.append({
                "keyword": m.group(1).strip().lower(),
                "target":  m.group(2).strip().upper(),
            })
    return results


# ── Main orchestration ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Upload Release Management Platform tickets to Jira Cloud."
    )
    parser.add_argument("--file", "-f", required=True,
                        help="Path to the .xlsx file")
    parser.add_argument("--dry-run", action="store_true",
                        help="Log all actions without making API calls")
    parser.add_argument("--skip-issues", action="store_true",
                        help="Skip issue creation (use when tickets already exist in Jira)")
    parser.add_argument("--skip-transitions", action="store_true",
                        help="Create issues in default state only (no transitions)")
    parser.add_argument("--skip-links", action="store_true",
                        help="Skip issue link creation")
    args = parser.parse_args()

    base_url = JIRA_BASE_URL
    email    = JIRA_USER_EMAIL
    token    = JIRA_API_TOKEN
    dry_run  = args.dry_run or DRY_RUN

    # Validate credentials
    if not base_url or not email or not token:
        log.error(
            "Missing Jira credentials.\n"
            "Ensure your .env file contains:\n"
            "  ATLASSIAN_DOMAIN=https://yourcompany.atlassian.net\n"
            "  ATLASSIAN_EMAIL=you@example.com\n"
            "  ATLASSIAN_API_TOKEN=your-api-token"
        )
        sys.exit(1)

    # Validate file
    if not Path(args.file).exists():
        log.error("File not found: %s", args.file)
        sys.exit(1)

    client = JiraClient(base_url, email, token, dry_run=dry_run)

    if dry_run:
        log.info("=== DRY-RUN MODE — no changes will be made ===")

    # ── Step 1: Read Excel ─────────────────────────────────────────────────────
    tickets  = read_jira_sheet(args.file)
    projects = sorted(set(t["Project"] for t in tickets if t.get("Project")))
    log.info("Projects in file: %s", projects)

    # ── Step 2: Verify projects exist in Jira ─────────────────────────────────
    log.info("\n--- Verifying Projects ---")
    for pkey in projects:
        proj = client.get_project(pkey)
        if proj is None and not dry_run:
            log.error(
                "Project '%s' does not exist in Jira.\n"
                "Create it first at: %s/jira/projects",
                pkey, base_url
            )
            sys.exit(1)
        log.info("Project %s: %s", pkey, proj.get("name", "OK") if proj else "dry-run")

    # ── Step 3: Ensure fix versions ────────────────────────────────────────────
    log.info("\n--- Creating Fix Versions ---")
    for pkey in projects:
        if pkey in FIX_VERSIONS:
            client.ensure_fix_version(pkey, FIX_VERSIONS[pkey])

    # ── Step 4: Create issues (or load existing) ──────────────────────────────
    if args.skip_issues:
        log.info("\n--- Loading Existing Issues (--skip-issues) ---")
        client.load_existing_keys(projects, tickets)
        created = list(client.excel_to_actual.keys())
        failed  = [t["Ticket Key"] for t in tickets
                   if t["Ticket Key"] not in client.excel_to_actual]
        log.info("Mapped: %d  |  Not found: %d", len(created), len(failed))
        if failed:
            log.warning("Could not map: %s", failed)
    else:
        log.info("\n--- Creating Issues ---")
        created, failed = [], []
        for ticket in tickets:
            key = client.create_issue(ticket.get("Project", ""), ticket)
            (created if key else failed).append(ticket["Ticket Key"])
        log.info("Created: %d  |  Failed: %d", len(created), len(failed))
        if failed:
            log.warning("Failed tickets: %s", failed)

    # ── Step 5: Transition to target statuses ──────────────────────────────────
    if not args.skip_transitions:
        log.info("\n--- Transitioning Issues ---")
        for ticket in tickets:
            excel_key  = ticket["Ticket Key"]
            actual_key = client.excel_to_actual.get(excel_key, excel_key)
            status     = ticket.get("Status", "")
            if actual_key in client.key_to_id and status:
                client.transition_issue(actual_key, status)
    else:
        log.info("\n--- Skipping transitions ---")

    # ── Step 6: Create issue links ─────────────────────────────────────────────
    if not args.skip_links:
        log.info("\n--- Creating Issue Links ---")
        for ticket in tickets:
            excel_source  = ticket["Ticket Key"]
            actual_source = client.excel_to_actual.get(excel_source, excel_source)
            for link in parse_links(ticket.get("Links", "")):
                kw           = link["keyword"]
                excel_target = link["target"]
                actual_target = client.excel_to_actual.get(excel_target, excel_target)
                defn         = LINK_TYPE_MAP.get(kw)
                if not defn:
                    log.warning("Unknown link keyword '%s' in %s (skipped)", kw, excel_source)
                    continue
                if defn["direction"] == "outward":
                    client.create_link(defn["type"], actual_target, actual_source)
                else:
                    client.create_link(defn["type"], actual_source, actual_target)
    else:
        log.info("\n--- Skipping links ---")

    # ── Summary ────────────────────────────────────────────────────────────────
    log.info("\n" + "=" * 60)
    log.info("UPLOAD COMPLETE")
    log.info("  Created : %d / %d tickets", len(created), len(tickets))
    log.info("  Failed  : %d", len(failed))
    log.info("  Dry-run : %s", dry_run)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
