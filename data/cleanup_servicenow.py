"""
Deletes all MVP ServiceNow records created by populate_servicenow.py.
Reads record identifiers directly from the Excel file.
Order: Change Tasks → Change Requests → CMDB CIs
"""

import os
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(override=True)

BASE_URL = os.getenv("SERVICENOW_INSTANCE_URL")
AUTH = (os.getenv("SERVICENOW_USERNAME"), os.getenv("SERVICENOW_PASSWORD"))
HEADERS = {"Content-Type": "application/json", "Accept": "application/json"}


def load_mvp_data(filepath: str) -> tuple[list, list, list]:
    """Read MVP short descriptions and CI names from Excel."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb["ServiceNow Data"]

    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
    mvp_rows = [r for r in rows if r[10] and "MVP" in str(r[10])]

    change_requests = [r[2] for r in mvp_rows if r[0] == "Change Request"]
    change_tasks    = [r[2] for r in mvp_rows if r[0] == "Change Task"]
    cmdb_cis        = [r[1] for r in mvp_rows if r[0] == "CMDB CI"]

    return change_requests, change_tasks, cmdb_cis


def get_sys_ids(table: str, field: str, values: list[str]) -> list[tuple[str, str]]:
    """Query a table and return (label, sys_id) pairs for matching records."""
    results = []
    for value in values:
        response = requests.get(
            f"{BASE_URL}/api/now/table/{table}",
            auth=AUTH,
            headers=HEADERS,
            params={"sysparm_query": f"{field}={value}", "sysparm_fields": "sys_id"},
        )
        response.raise_for_status()
        for record in response.json().get("result", []):
            results.append((value, record["sys_id"]))
    return results


def delete_record(table: str, label: str, sys_id: str):
    """Delete a single record by sys_id."""
    response = requests.delete(
        f"{BASE_URL}/api/now/table/{table}/{sys_id}",
        auth=AUTH,
        headers=HEADERS,
    )
    response.raise_for_status()
    print(f"  ✓ Deleted {table}: {label}")


def delete_all(table: str, field: str, values: list[str]):
    """Find and delete all matching records in a table."""
    records = get_sys_ids(table, field, values)
    if not records:
        print(f"  — No records found in {table}")
        return
    for label, sys_id in records:
        delete_record(table, label, sys_id)


def main():
    filepath = "data_loaders/Release_Management_Platform_Sample_Data.xlsx"
    change_requests, change_tasks, cmdb_cis = load_mvp_data(filepath)

    # Order matters: tasks before requests, requests before CIs
    print("\n🗑️  Deleting Change Tasks...")
    delete_all("change_task", "short_description", change_tasks)

    print("\n🗑️  Deleting Change Requests...")
    delete_all("change_request", "short_description", change_requests)

    print("\n🗑️  Deleting CMDB Configuration Items...")
    delete_all("cmdb_ci", "name", cmdb_cis)

    print("\n✅ Cleanup complete!")


if __name__ == "__main__":
    main()