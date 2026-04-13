"""
Populates Confluence with MVP data from the Excel file and .md sample content.
Structure:
    Release Management (RM) space
    └── ATLAS
        └── ATLAS v3.2 — Release Plan
        └── ATLAS v3.2 — Deployment Runbook
        └── ...
    └── BEACON
    └── CIPHER
    └── DELTA
    └── Release Runbook Template

Safe to re-run — skips pages and spaces that already exist.
Space is created via Confluence REST API v1 directly (atlassian-python-api
uses v2 which requires RBAC Early Access Program enrollment).
"""

import os
import requests
from requests.auth import HTTPBasicAuth
from atlassian import Confluence
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(override=True)

SPACE_KEY  = "RMP"
SPACE_NAME = "Release Management Platform"
EXCEL_PATH = "data_loaders/Release_Management_Platform_Sample_Data.xlsx"

BASE_URL = os.getenv("JIRA_BASE_URL")
AUTH     = HTTPBasicAuth(os.getenv("JIRA_EMAIL"), os.getenv("JIRA_API_TOKEN"))
HEADERS  = {"Content-Type": "application/json", "Accept": "application/json"}


def get_client() -> Confluence:
    return Confluence(
        url=BASE_URL.rstrip("/"),
        username=os.getenv("JIRA_EMAIL"),
        password=os.getenv("JIRA_API_TOKEN"),
        cloud=True,
    )


def load_mvp_pages(filepath: str) -> list[dict]:
    """Read MVP Confluence pages from Excel. Returns only pages that should exist."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb["Confluence Pages"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
    return [
        {"project": r[0], "title": r[1], "quality": r[3], "notes": r[5]}
        for r in rows
        if r[6] == "MVP" and r[2] == "Yes"
    ]


def space_exists() -> bool:
    """Check if RM space exists using REST API directly."""
    response = requests.get(
        f"{BASE_URL.rstrip('/')}/wiki/rest/api/space/{SPACE_KEY}",
        auth=AUTH,
        headers=HEADERS,
    )
    return response.status_code == 200


def create_space_via_api():
    """Create RM space using Confluence REST API v1 directly."""
    response = requests.post(
        f"{BASE_URL.rstrip('/')}/wiki/rest/api/space",
        auth=AUTH,
        headers=HEADERS,
        json={
            "key": SPACE_KEY,
            "name": SPACE_NAME,
            "description": {
                "plain": {
                    "value": "Release Management space for tracking releases across all projects.",
                    "representation": "plain"
                }
            }
        },
    )
    if not response.ok:
        print(f"  ✗ Error: {response.text}")
    response.raise_for_status()
    print(f"  ✓ Space created")


def create_page_if_not_exists(
    client: Confluence, space: str, title: str, body: str, parent_id: str = None
) -> dict:
    """Create a page only if it doesn't already exist. Returns the page."""
    existing = client.get_page_by_title(space=space, title=title)
    if existing:
        print(f"  — Already exists: {title}")
        return existing
    kwargs = {"space": space, "title": title, "body": body}
    if parent_id:
        kwargs["parent_id"] = parent_id
    page = client.create_page(**kwargs)
    print(f"  ✓ Created: {title} — {page['id']}")
    return page


# ── Page content generators ───────────────────────────────────────────────────

def release_plan_content(project: str, quality: str) -> str:
    if project == "ATLAS":
        return """<h2>Scope Summary</h2>
<p>Payment processing microservice upgrade to v3.2 including PCI compliance tokenization,
schema migration, and gateway API integration update.</p>
<h2>Key Dates</h2>
<ul>
<li>Code freeze: May 8, 2026</li>
<li>QA sign-off deadline: May 13, 2026</li>
<li>CAB submission: May 6, 2026</li>
<li>Go/No-Go: May 14, 2026</li>
<li>Deployment: May 16, 2026 (22:00–02:00 UTC)</li>
</ul>
<h2>Team Contacts</h2>
<ul>
<li>Engineering Lead: Alice Chen (alice.chen@company.com)</li>
<li>QA Lead: Bob Martinez (bob.martinez@company.com)</li>
<li>DevOps: Carol Singh (carol.singh@company.com)</li>
</ul>
<h2>Artifacts</h2>
<ul>
<li>atlas-payment-service:3.2.0 (Docker image)</li>
<li>atlas-db-migration:3.2.0 (SQL scripts)</li>
</ul>
<h2>Environment Schedule</h2>
<ul>
<li>Pre-Prod deployment: May 10, 2026</li>
<li>UAT window: May 11–13, 2026</li>
</ul>"""

    if project == "BEACON":
        return """<h2>Scope Summary</h2>
<p>Customer portal complete redesign to v2.0 including responsive mobile view,
dark mode support, and new component library migration.</p>
<h2>Key Dates</h2>
<ul>
<li>Code freeze: May 9, 2026</li>
<li>QA sign-off deadline: May 13, 2026</li>
<li>CAB submission: May 7, 2026</li>
<li>Go/No-Go: May 14, 2026</li>
<li>Deployment: May 16, 2026 (22:00–00:00 UTC)</li>
</ul>
<h2>Team Contacts</h2>
<ul>
<li>Engineering Lead: Dev5 (dev5@company.com)</li>
<li>QA Lead: Dev7 (dev7@company.com)</li>
</ul>
<h2>Dependencies</h2>
<p>BEACON-105 is blocked by ATLAS-102. Deployment depends on Atlas releasing first.</p>"""

    if project == "CIPHER":
        return """<h2>Scope Summary</h2>
<p>API gateway security hardening including WAF rules, mTLS authentication,
and rate limiting on public endpoints.</p>
<ac:structured-macro ac:name="warning">
  <ac:parameter ac:name="title">Incomplete</ac:parameter>
  <ac:rich-text-body><p>Key dates and environment schedule are missing. This page requires updating.</p></ac:rich-text-body>
</ac:structured-macro>
<h2>Team Contacts</h2>
<ul>
<li>Engineering Lead: Dev8 (dev8@company.com)</li>
<li>QA Lead: Dev9 (dev9@company.com)</li>
</ul>"""

    if project == "DELTA":
        return """<h2>Scope Summary</h2>
<p>Internal reporting dashboard enhancements including PDF export,
date range filters, and chart type selector.</p>
<h2>Key Dates</h2>
<ul>
<li>Code freeze: May 8, 2026</li>
<li>QA sign-off deadline: May 13, 2026</li>
<li>CAB submission: May 6, 2026</li>
<li>Go/No-Go: May 14, 2026</li>
<li>Deployment: May 16, 2026 (22:00–23:00 UTC)</li>
</ul>
<h2>Team Contacts</h2>
<ul>
<li>Engineering Lead: Dev10 (dev10@company.com)</li>
<li>QA Lead: Dev11 (dev11@company.com)</li>
</ul>"""


def runbook_content(project: str, quality: str) -> str:
    base = """<h2>Deployment Steps</h2>
<table>
<tr><th>Step</th><th>Action</th><th>Owner</th><th>Duration</th><th>Verification</th></tr>"""

    if project == "ATLAS":
        return base + """
<tr><td>1</td><td>Run atlas-db-migration:3.2.0 against payment_transactions database</td>
<td>Carol Singh</td><td>15 min</td><td>Check migration_log table for successful completion</td></tr>
<tr><td>2</td><td>Deploy atlas-payment-service:3.2.0 to production K8s cluster (rolling deployment)</td>
<td>Carol Singh</td><td>10 min</td><td>Hit /health endpoint — expect 200 OK</td></tr>
<tr><td>3</td><td>Execute smoke tests — process test transaction</td>
<td>Bob Martinez</td><td>5 min</td><td>Expect success response</td></tr>
</table>
<h2>Rollback Steps</h2>
<p><em>(TODO — to be completed)</em></p>
<h2>Verification Checks</h2>
<p><em>(TODO — to be completed)</em></p>"""

    steps = {
        "BEACON": [
            ("1", "Deploy customer-portal:2.0.0 to production", "Dev5", "10 min", "Check /health endpoint"),
            ("2", "Run smoke tests — cross browser validation", "Dev7", "10 min", "All tests pass"),
            ("3", "Validate dark mode rendering", "Dev5", "5 min", "No visual regressions"),
        ],
        "DELTA": [
            ("1", "Deploy delta-reporting:4.1.0 to reporting-server-prod", "Dev11", "5 min", "Check /health endpoint"),
            ("2", "Run regression tests", "Dev10", "10 min", "All tests pass"),
            ("3", "Validate PDF export functionality", "Dev10", "5 min", "PDF generates correctly"),
        ],
    }
    rows = "".join(
        f"<tr><td>{s}</td><td>{a}</td><td>{o}</td><td>{d}</td><td>{v}</td></tr>"
        for s, a, o, d, v in steps.get(project, [])
    )
    return base + rows + """</table>
<h2>Rollback Steps</h2>
<ol>
<li>Redeploy previous version from registry</li>
<li>Verify /health endpoint returns 200</li>
<li>Notify stakeholders of rollback</li>
</ol>
<h2>Verification Checks</h2>
<ol>
<li>All health endpoints return 200</li>
<li>Key user journeys functional</li>
<li>No error spike in monitoring dashboard</li>
</ol>
<h2>Contacts</h2>
<ul>
<li>On-call DevOps: devops-oncall@company.com</li>
<li>Escalation: engineering-lead@company.com</li>
</ul>"""


def test_evidence_content(project: str) -> str:
    results = {
        "ATLAS": [
            ("ATLAS-201", "Test PCI tokenization — functional",    "Passed"),
            ("ATLAS-202", "Test schema migration — data integrity", "Passed"),
            ("ATLAS-203", "Test gateway API — integration",         "Not Started"),
            ("ATLAS-204", "Test retry logic — edge cases",          "Passed"),
        ],
        "BEACON": [
            ("BEACON-201", "Test dashboard layout — cross-browser", "Passed"),
            ("BEACON-202", "Test mobile responsiveness",            "Passed"),
            ("BEACON-203", "Test dark mode — accessibility",        "Passed"),
            ("BEACON-204", "Test component library migration",      "Passed"),
        ],
        "DELTA": [
            ("DELTA-201", "Test PDF export — formatting",           "Passed"),
            ("DELTA-202", "Test date filter — edge cases",          "Passed"),
            ("DELTA-203", "Test chart selector — rendering",        "Passed"),
        ],
    }
    rows = "".join(
        f"<tr><td>{k}</td><td>{s}</td><td>{r}</td></tr>"
        for k, s, r in results.get(project, [])
    )
    return f"""<h2>Test Results Summary</h2>
<table>
<tr><th>Ticket</th><th>Test</th><th>Result</th></tr>
{rows}
</table>"""


def rollback_content(project: str) -> str:
    return f"""<h2>Rollback Trigger Criteria</h2>
<ul>
<li>Health endpoint returns non-200 after deployment</li>
<li>Error rate exceeds 1% within 15 minutes of deployment</li>
<li>Critical business function unavailable</li>
</ul>
<h2>Rollback Steps</h2>
<ol>
<li>Notify on-call team and stakeholders immediately</li>
<li>Redeploy previous stable version from container registry</li>
<li>Verify all health checks pass post-rollback</li>
<li>Raise ServiceNow incident linked to the Change Request</li>
<li>Document rollback reason in post-release report</li>
</ol>
<h2>Contacts</h2>
<ul>
<li>On-call: devops-oncall@company.com</li>
<li>Engineering Lead: {project.lower()}-lead@company.com</li>
</ul>"""


def release_notes_content(project: str, quality: str) -> str:
    notes = {
        "ATLAS": """<ac:structured-macro ac:name="info">
  <ac:parameter ac:name="title">Work In Progress</ac:parameter>
  <ac:rich-text-body><p>These release notes are a draft and subject to change.</p></ac:rich-text-body>
</ac:structured-macro>
<h2>What's New in v3.2</h2>
<ul>
<li>PCI-compliant tokenization for card data</li>
<li>Updated payment gateway API integration</li>
<li>Retry logic for failed payment callbacks</li>
<li>Audit logging for payment events</li>
</ul>
<h2>Known Issues</h2>
<p><em>(To be completed before release)</em></p>""",
        "BEACON": """<h2>What's New in v2.0</h2>
<ul>
<li>Fully redesigned customer dashboard layout</li>
<li>Responsive mobile view</li>
<li>Dark mode theme support</li>
<li>Migrated to new component library</li>
</ul>
<h2>Known Issues</h2>
<ul><li>Minor CSS rendering issue in Safari — fix scheduled for next release</li></ul>""",
        "DELTA": """<h2>What's New in v4.1</h2>
<ul>
<li>Export reports to PDF</li>
<li>Date range filter on dashboards</li>
<li>Chart type selector (bar / line / pie)</li>
</ul>
<h2>Known Issues</h2>
<p>None.</p>""",
    }
    return notes.get(project, f"<p>Release notes for {project}.</p>")


def runbook_template_content() -> str:
    return """<h2>Purpose</h2>
<p>This template defines the required structure for all deployment runbooks.
All sections must be completed before a release can be approved at Go/No-Go.</p>
<h2>Required Sections</h2>
<h3>1. Deployment Steps</h3>
<p>Each step must include: action, owner, duration estimate, verification check.</p>
<h3>2. Rollback Steps</h3>
<p>Each step must include: trigger criteria, action, owner, duration estimate.</p>
<h3>3. Verification Checks</h3>
<p>Post-deployment validation steps to confirm the release is stable.</p>
<h3>4. Contacts</h3>
<p>Deployment team, escalation path, on-call contact.</p>
<h3>5. Prerequisites</h3>
<p>Infrastructure readiness, certificates, secrets, approvals required before deployment.</p>
<ac:structured-macro ac:name="warning">
  <ac:parameter ac:name="title">Validation</ac:parameter>
  <ac:rich-text-body>
    <p>The Release Manager Agent will validate this runbook structure before Go/No-Go.
    Missing sections will block release approval.</p>
  </ac:rich-text-body>
</ac:structured-macro>"""


def get_page_content(title: str, project: str, quality: str) -> str:
    """Return appropriate content based on page type."""
    t = title.lower()
    if "release plan"       in t: return release_plan_content(project, quality)
    if "deployment runbook" in t: return runbook_content(project, quality)
    if "test evidence"      in t: return test_evidence_content(project)
    if "rollback procedure" in t: return rollback_content(project)
    if "release notes"      in t: return release_notes_content(project, quality)
    if "runbook template"   in t: return runbook_template_content()
    return f"<p>{title}</p>"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    client = get_client()
    pages  = load_mvp_pages(EXCEL_PATH)

    # Step 1 — Create RM space via REST API v1 if it doesn't exist
    print(f"\n📁 Creating space: {SPACE_NAME} ({SPACE_KEY})...")
    if space_exists():
        print(f"  — Space already exists")
    else:
        create_space_via_api()

    # Step 2 — Create project parent pages, track their IDs
    projects = ["ATLAS", "BEACON", "CIPHER", "DELTA"]
    parent_ids = {}
    print("\n📁 Creating project parent pages...")
    for project in projects:
        page = create_page_if_not_exists(
            client, SPACE_KEY, project,
            f"<p>Release documentation for {project}.</p>",
        )
        parent_ids[project] = page["id"]

    # Step 3 — Create child pages under each project parent
    print("\n📄 Creating project pages...")
    for p in pages:
        project   = p["project"]
        title     = p["title"]
        quality   = p["quality"]
        parent_id = parent_ids.get(project)
        content   = get_page_content(title, project, quality)
        create_page_if_not_exists(client, SPACE_KEY, title, content, parent_id)

    # Step 4 — Create Runbook Template at root level
    print("\n📄 Creating Runbook Template...")
    create_page_if_not_exists(
        client, SPACE_KEY,
        "Release Runbook Template",
        runbook_template_content(),
    )

    print("\n✅ MVP Confluence data population complete!")


if __name__ == "__main__":
    main()