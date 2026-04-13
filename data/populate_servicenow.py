"""
Populates ServiceNow with MVP data read directly from the Excel file.
Creates: CMDB CIs, Change Requests (with all CI names in description), and Change Tasks.
"""

import os
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(override=True)

BASE_URL = os.getenv("SERVICENOW_INSTANCE_URL")
AUTH = (os.getenv("SERVICENOW_USERNAME"), os.getenv("SERVICENOW_PASSWORD"))
HEADERS = {"Content-Type": "application/json", "Accept": "application/json"}

RISK_MAP = {"High": "3", "Medium": "2", "Low": "1"}

PROJECT_TO_CR = {
    "Atlas": "ATLAS v3.2 — Payment service upgrade",
    "Delta": "DELTA v4.1 — Reporting dashboard enhancements",
}


def load_mvp_data(filepath: str) -> tuple[list, list, list]:
    """Read MVP Change Requests, Change Tasks, and CMDB CIs from Excel."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb["ServiceNow Data"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
    mvp_rows = [r for r in rows if r[10] and "MVP" in str(r[10])]
    return (
        [r for r in mvp_rows if r[0] == "Change Request"],
        [r for r in mvp_rows if r[0] == "Change Task"],
        [r for r in mvp_rows if r[0] == "CMDB CI"],
    )


def create_record(table: str, payload: dict) -> dict:
    """Create a record in a ServiceNow table and return it."""
    response = requests.post(
        f"{BASE_URL}/api/now/table/{table}",
        auth=AUTH,
        headers=HEADERS,
        json=payload,
    )
    response.raise_for_status()
    result = response.json()["result"]
    print(f"  ✓ Created {table}: {result.get('number') or result.get('name')} — {result['sys_id']}")
    return result


def get_sys_id(table: str, field: str, value: str) -> str:
    """Look up a record's sys_id by field value. Returns empty string if not found."""
    response = requests.get(
        f"{BASE_URL}/api/now/table/{table}",
        auth=AUTH,
        headers=HEADERS,
        params={"sysparm_query": f"{field}={value}", "sysparm_fields": "sys_id"},
    )
    response.raise_for_status()
    results = response.json().get("result", [])
    return results[0]["sys_id"] if results else ""


def main():
    filepath = "data_loaders/Release_Management_Platform_Sample_Data.xlsx"
    change_requests, change_tasks, cmdb_cis = load_mvp_data(filepath)

    # Step 1 — Create CMDB CIs, track shared ones (project = "All")
    print("\n📋 Creating CMDB Configuration Items...")
    shared_cis = []
    for _, name, short_desc, project, *_ in cmdb_cis:
        create_record("cmdb_ci", {"name": name, "short_description": short_desc})
        if str(project).strip().lower() == "all":
            shared_cis.append(name)

    # Step 2 — Create Change Requests
    print("\n📋 Creating Change Requests...")
    cr_map = {}  # short_description -> sys_id
    for _, _, short_desc, description, _, risk, _, start, end, cmdb_ci, _ in change_requests:
        # Build full CI list: CR's own CIs + shared CIs
        ci_names = [c.strip() for c in cmdb_ci.split(",") if c.strip()] if cmdb_ci else []
        all_ci_names = ci_names + [ci for ci in shared_cis if ci not in ci_names]

        # Look up sys_id for primary CI
        primary_ci_sys_id = get_sys_id("cmdb_ci", "name", ci_names[0]) if ci_names else ""

        # Build description with team info and full CI list
        full_description = f"{description}\n\nAffected CIs:\n" + "\n".join(f"- {ci}" for ci in all_ci_names)

        result = create_record("change_request", {
            "short_description": short_desc,
            "description": full_description,
            "risk": RISK_MAP.get(risk, "2"),
            "start_date": str(start),
            "end_date": str(end),
            "cmdb_ci": primary_ci_sys_id,
            "type": "normal",
            "category": "Software",
        })
        cr_map[short_desc] = result["sys_id"]

    # Step 3 — Create Change Tasks linked to their parent CRs
    print("\n📋 Creating Change Tasks...")
    for _, _, short_desc, project, *_ in change_tasks:
        parent_desc = PROJECT_TO_CR.get(project)
        create_record("change_task", {
            "short_description": short_desc,
            "change_request": cr_map.get(parent_desc, ""),
        })

    print("\n✅ MVP ServiceNow data population complete!")


if __name__ == "__main__":
    main()